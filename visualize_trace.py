import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import argparse
import random

parser = argparse.ArgumentParser(description="pipeline and model chunks")
parser.add_argument("--pipeline_parallel", required=True, help="pipeline parallel size")
parser.add_argument("--model_chunks", required=True, help="number of model chunks")
args = parser.parse_args()

num_gpus = 4

center_alignment = Alignment(horizontal="center", vertical="center")

pipeline_parallel = int(args.pipeline_parallel)
data_parallel = num_gpus // int(args.pipeline_parallel)
num_microbatches = 16
num_model_chunks = int(args.model_chunks)
sample = random.randint(0, 19)

events = []

file_name = f"trace_pipeline_{pipeline_parallel}_model_chunks_{num_model_chunks}.json"

with open(file_name) as file:
    events = json.load(file)


def get_color(depth, direction):
    color_intensity = min(255, max(0, 255 - depth * 40))

    if direction == "forward":
        return PatternFill(
            start_color=f"0080{color_intensity:02X}",
            end_color=f"0080{color_intensity:02X}",
            fill_type="solid",
        )
    else:
        return PatternFill(
            start_color=f"00{color_intensity:02X}00",
            end_color=f"00{color_intensity:02X}00",
            fill_type="solid",
        )


wb = Workbook()
ws = wb.active


def number_to_column(n):
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


for pipeline_rank in range(pipeline_parallel - 1, -1, -1):
    filtered_events = [
        event
        for event in events
        if event["pid"] == pipeline_rank * data_parallel
        and event["ph"] == "B"
        and (
            event["name"].startswith("forward") or event["name"].startswith("backward")
        )
    ]
    sorted_events = sorted(filtered_events, key=lambda x: x["ts"])[
        sample
        * num_microbatches
        * num_model_chunks
        * 2 : (sample + 1)
        * num_microbatches
        * num_model_chunks
        * 2
    ]
    cur_row = pipeline_rank + 1
    cur_index = pipeline_rank + 1
    for event in sorted_events:
        name = event["name"]
        a, b = name.split(" ")[1].split("-")
        a = int(a)
        b = int(b)
        if "forward" in name:
            direction = "forward"
        else:
            direction = "backward"
            a = num_model_chunks - 1 - a

        formatted_value = f"{a}-{b}"
        count = 0
        if direction == "backward" and cur_row < pipeline_parallel:
            i = 1
            while True:
                cur_value = ws.cell(row=cur_row + 1, column=i).value
                if cur_value == formatted_value:
                    count += 1
                    if count == 2:
                        cur_index = max(i + 2, cur_index)
                        break
                i += 1
        fill = get_color(a, direction)
        cell = ws.cell(
            row=cur_row,
            column=cur_index,
            value=formatted_value,
        )
        cell.fill = fill
        cell.alignment = center_alignment
        if direction == "backward":
            ws.merge_cells(
                start_row=cur_row,
                start_column=cur_index,
                end_row=cur_row,
                end_column=cur_index + 1,
            )
            cur_index += 2
        else:
            cur_index += 1

for col in range(1, cur_index + 1):
    col_letter = number_to_column(col)
    ws.column_dimensions[col_letter].width = 4

wb.save(f"{file_name[:-5]}.xlsx")
